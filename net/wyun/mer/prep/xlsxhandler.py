from net.wyun.mer.ink import scginkparser
import numpy as np
from net.wyun.mer.ink.stroke import Stroke
from net.wyun.mer.ink.stroke import get_bounding_box
from net.wyun.mer.ink.stroke import get_bounding_box_h1000
from net.wyun.mer.ink.sample import Sample
from scipy import misc

# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('data/scg/test.xlsx')

# Get sheet names
print(wb.get_sheet_names())  #[u'hw_record', u'Sheet1']

# Get a sheet by name
ws = wb.get_sheet_by_name('hw_record')

# Print the sheet title
print ws['A1'].value, ws['B1'].value, ws['C1'].value

print ws['A2'].value, ws['B2'].value, ws['C2'].value

print 'length of record: ', len(ws['A'])


scg_id = int(ws['A2'].value)
scg_content = ws['B2'].value

strokes = scginkparser.parse_scg_ink_file(scg_content, scg_id)

for st in strokes:
    print st

traces = {}

trace_id_int = 0
for st in strokes:
    coords = np.zeros((2, len(st)))
    idx = 0
    for x_y in st:
        coords[:, idx] = [float(x_y[0]), float(x_y[1])]
        idx += 1
    traces[trace_id_int] = Stroke(trace_id_int, coords)
    trace_id_int += 1


# //Compute bounding box of the input expression
x_min, y_min, x_max, y_max = get_bounding_box(traces)  # bounding box for the whole math expression

# Just in case there is only one point or a sequence of	points perfectly aligned with the x or y axis
if x_max == x_min: x_max = x_min + 1;
if y_max == y_min: y_max = y_min + 1;

# Renormalize to height [0,10000] keeping the aspect ratio
H = 10000.0
W = H * (x_max - x_min) / (y_max - y_min)
for trace_key, trace_v in traces.iteritems():
    trace_v.calc_coords_h10000(H, W, x_min, y_min, x_max, y_max)


for trace_key, trace_v in traces.iteritems():
    print trace_key, trace_v
    rx, ry, rs, rt = trace_v.get_bounding_box_h10000()
    print rx, ry, rs, rt

dummy_sample =  Sample('data/inkml/65_alfonso.inkml')
dummy_sample.traces = traces
img, W, H = dummy_sample.render()
print 'save image to temp/all.png: '
misc.imsave('temp/all.png', img)
